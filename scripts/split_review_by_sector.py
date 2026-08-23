#!/usr/bin/env python
"""Split a review xlsx into one sheet per sector.

Reads sector assignments from technologies.xlsx and splits the input review
file (review.xlsx or any variant like review_with-manual-compare.xlsx) into
a multi-sheet workbook with one sheet per sector plus an "All" sheet.

All formatting (fills, fonts, hyperlinks, column widths, grouping, freeze
panes, auto-filter) is faithfully copied.

Usage:
    conda run -n env-brla python scripts/split_review_by_sector.py \
        --review output/review.xlsx

    conda run -n env-brla python scripts/split_review_by_sector.py \
        --review output/review_with-manual-compare.xlsx

    # Custom technologies file or output path:
    conda run -n env-brla python scripts/split_review_by_sector.py \
        --review output/review.xlsx \
        --technologies input/technologies.xlsx \
        --output output/review-sector-wise.xlsx

    python scripts/split_review_by_sector.py \
        --review output/review_with-manual-compare.xlsx \
        --technologies input/technologies.xlsx \
        --output output/review_with-manual-compare-sector-wise.xlsx
"""
import argparse
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def load_sector_map(tech_path: Path) -> dict[str, str]:
    """Return {tech_id: sector} from technologies.xlsx."""
    wb = load_workbook(tech_path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}

    if "tech_id" not in col:
        raise RuntimeError(
            f"technologies.xlsx missing 'tech_id' column. Found: {headers}")
    if "sector" not in col:
        raise RuntimeError(
            f"technologies.xlsx missing 'sector' column. Found: {headers}")

    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        tid = row[col["tech_id"]]
        sector = row[col["sector"]]
        if tid:
            result[str(tid).strip()] = str(sector).strip() if sector else ""
    wb.close()
    return result


def _copy_cell(src_cell, dst_cell):
    """Copy value, style, and hyperlink from one cell to another."""
    dst_cell.value = src_cell.value
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.border = copy(src_cell.border)
    if src_cell.hyperlink:
        dst_cell.hyperlink = src_cell.hyperlink


def _write_sheet(ws_dst, header_row, data_rows, col_widths, grouped_range,
                 sheet_name):
    """Write a fully formatted sheet from pre-read row data."""
    safe_name = sheet_name.translate(str.maketrans(r'/\?*[]:', '-------'))[:31]
    ws_dst.title = safe_name
    ws_dst.freeze_panes = "A2"
    n_cols = len(header_row)

    for col_idx, src_cell in enumerate(header_row, 1):
        dst = ws_dst.cell(row=1, column=col_idx)
        _copy_cell(src_cell, dst)

    for row_offset, src_row in enumerate(data_rows, 2):
        for col_idx, src_cell in enumerate(src_row, 1):
            dst = ws_dst.cell(row=row_offset, column=col_idx)
            _copy_cell(src_cell, dst)

    if grouped_range:
        ws_dst.column_dimensions.group(
            grouped_range[0], grouped_range[1], hidden=False)

    for col_idx in range(1, n_cols + 1):
        letter = get_column_letter(col_idx)
        if letter in col_widths:
            ws_dst.column_dimensions[letter].width = col_widths[letter]

    total_rows = len(data_rows) + 1
    ws_dst.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{total_rows}"


def main():
    p = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--review", default="output/review.xlsx",
                   help="input review xlsx (default: output/review.xlsx)")
    p.add_argument("--technologies", default="input/technologies.xlsx",
                   help="technologies xlsx with sector column "
                        "(default: input/technologies.xlsx)")
    p.add_argument("--output", default=None,
                   help="output path (default: review-sector-wise.xlsx next "
                        "to input)")
    args = p.parse_args()

    review_path = Path(args.review)
    tech_path = Path(args.technologies)

    if not review_path.exists():
        raise FileNotFoundError(f"Review file not found: {review_path}")
    if not tech_path.exists():
        raise FileNotFoundError(f"Technologies file not found: {tech_path}")

    if args.output:
        output_path = Path(args.output)
    else:
        output_path = review_path.parent / (
            review_path.stem + "-sector-wise.xlsx")

    sector_map = load_sector_map(tech_path)
    print(f"Loaded {len(sector_map)} tech→sector mappings from {tech_path}")

    src_wb = load_workbook(review_path)
    src_ws = src_wb.active
    n_cols = src_ws.max_column

    headers = [c.value for c in src_ws[1]]
    if "tech_id" not in headers:
        raise RuntimeError(
            f"Review xlsx missing 'tech_id' column. Found: {headers}")
    tid_col_idx = headers.index("tech_id")

    header_cells = list(src_ws[1])[:n_cols]

    col_widths = {}
    for col_idx in range(1, n_cols + 1):
        letter = get_column_letter(col_idx)
        dim = src_ws.column_dimensions.get(letter)
        if dim and dim.width:
            col_widths[letter] = dim.width

    source_letters = [get_column_letter(i + 1) for i, h in enumerate(headers)
                      if h and str(h).startswith("source_")]
    grouped_range = None
    if source_letters:
        grouped_range = [source_letters[0], source_letters[-1]]
        first_w = col_widths.get(source_letters[0])
        if first_w:
            for letter in source_letters[1:]:
                col_widths.setdefault(letter, first_w)

    all_rows = []
    for row_idx in range(2, src_ws.max_row + 1):
        row_cells = [src_ws.cell(row=row_idx, column=c)
                     for c in range(1, n_cols + 1)]
        all_rows.append(row_cells)

    sector_rows: dict[str, list] = {}
    unmapped = set()
    for row_cells in all_rows:
        tid = str(row_cells[tid_col_idx].value or "").strip()
        sector = sector_map.get(tid, "")
        if not sector:
            sector = "Uncategorized"
            if tid:
                unmapped.add(tid)
        sector_rows.setdefault(sector, []).append(row_cells)

    if unmapped:
        print(f"  Warning: {len(unmapped)} tech_ids have no sector: "
              f"{sorted(unmapped)}")

    dst_wb = Workbook()
    dst_wb.remove(dst_wb.active)

    ws_all = dst_wb.create_sheet()
    _write_sheet(ws_all, header_cells, all_rows, col_widths, grouped_range,
                 "All")

    for sector in sorted(sector_rows):
        ws = dst_wb.create_sheet()
        rows = sector_rows[sector]
        _write_sheet(ws, header_cells, rows, col_widths, grouped_range,
                     sector)
        print(f"  {sector}: {len(rows)} rows")

    dst_wb.save(output_path)
    print(f"\nWrote {output_path}")
    print(f"  {len(sector_rows)} sector sheets + All "
          f"({len(all_rows)} total rows)")


if __name__ == "__main__":
    main()
