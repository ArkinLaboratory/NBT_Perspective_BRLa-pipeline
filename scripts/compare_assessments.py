#!/usr/bin/env python
"""Compare pipeline vs manual BRLa assessments and generate an HTML dashboard.

Reads two Excel files — the pipeline's review.xlsx and a manually-performed BRLa
draft — and produces a self-contained HTML page showing agreement/disagreement
patterns with auto-generated insight text.

Usage:
    conda run -n env-brla python scripts/compare_assessments.py \
        --manual notes/assets/sheets/brla_livestock_draft-1.xlsx
"""

import argparse
import dataclasses
import json
import sys
from collections import Counter, defaultdict
from difflib import SequenceMatcher
from pathlib import Path
from string import Template

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DIMENSIONS = ["TRL", "MRL", "RRL", "ARL", "ORL"]
BIN_ORDER = {"High": 2, "Mid": 1, "Low": 0}


# ── Data loading ────────────────────────────────────────────────────────────


def load_pipeline(path: Path) -> dict[str, dict]:
    """Load pipeline review.xlsx into {tech_id: {tech_name, dims: {dim: {...}}}}."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}

    required = {"tech_id", "tech_name", "dimension", "bin"}
    missing = required - set(col)
    if missing:
        raise RuntimeError(f"review.xlsx missing columns: {missing}")

    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        tid = row[col["tech_id"]]
        dim = row[col["dimension"]]
        if not tid or not dim:
            continue
        tid = str(tid).strip()
        dim = str(dim).strip()

        if tid not in result:
            result[tid] = {"tech_name": str(row[col["tech_name"]]).strip(), "dims": {}}

        def _get(name, default=None):
            return row[col[name]] if name in col else default

        def _bool(name):
            v = _get(name)
            if v is None:
                return False
            if isinstance(v, bool):
                return v
            return str(v).strip().lower() in ("yes", "y", "true", "1")

        r2 = _get("rater_2_bin")
        result[tid]["dims"][dim] = {
            "bin": str(row[col["bin"]]).strip() if row[col["bin"]] else None,
            "r2_bin": str(r2).strip() if r2 else None,
            "agreement": _get("agreement"),
            "evidence_gap": _bool("evidence_gap"),
            "needs_review": _bool("needs_review"),
            "evidence_strength": _get("evidence_strength"),
            "n_sources": _get("n_sources"),
            "single_rater": _bool("single_rater"),
            "level_est": _get("level_est"),
        }

    wb.close()
    return result


def load_manual(path: Path) -> dict[str, dict[str, str]]:
    """Load manual xlsx into {tech_name: {TRL: bin, MRL: bin, ...}}.

    Skips rows where column 2 is None (continuation rows from merged cells).
    """
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        name = vals[1]
        if name is None:
            continue
        name = str(name).strip()
        bins = {}
        for i, dim in enumerate(DIMENSIONS):
            v = vals[2 + i] if (2 + i) < len(vals) else None
            bins[dim] = str(v).strip() if v else None
        result[name] = bins

    wb.close()
    return result


# ── Tech name matching ──────────────────────────────────────────────────────


def match_techs(
    manual_names: list[str],
    pipeline: dict[str, dict],
    mapping_path: Path | None,
    threshold: float,
) -> dict[str, str]:
    """Fuzzy-match manual tech names to pipeline tech_ids.

    Returns {manual_name: tech_id}. Prints matching table to stderr.
    """
    overrides = {}
    if mapping_path and mapping_path.exists():
        overrides = json.loads(mapping_path.read_text())

    pipe_names = {tid: d["tech_name"] for tid, d in pipeline.items()}
    matched = {}

    print("\n  Tech name matching:", file=sys.stderr)
    print(f"  {'Manual name':<50} {'Pipeline match':<45} {'Ratio':<6}", file=sys.stderr)
    print("  " + "-" * 101, file=sys.stderr)

    for mname in manual_names:
        if mname in overrides:
            tid = overrides[mname]
            if tid in pipeline:
                matched[mname] = tid
                print(f"  {mname:<50} {tid:<45} {'OVERRIDE':<6}", file=sys.stderr)
            else:
                print(f"  {mname:<50} {tid:<45} {'BAD ID':<6}", file=sys.stderr)
            continue

        best_tid = None
        best_ratio = 0.0
        for tid, pname in pipe_names.items():
            ratio = SequenceMatcher(None, mname.lower(), pname.lower()).ratio()
            if ratio > best_ratio:
                best_ratio = ratio
                best_tid = tid

        if best_ratio >= threshold and best_tid:
            matched[mname] = best_tid
            label = pipe_names[best_tid]
            print(f"  {mname:<50} {label:<45} {best_ratio:.3f}", file=sys.stderr)
        else:
            label = pipe_names.get(best_tid, "???") if best_tid else "???"
            print(
                f"  {mname:<50} {label:<45} {best_ratio:.3f}  ** BELOW THRESHOLD **",
                file=sys.stderr,
            )

    n = len(matched)
    print(f"\n  Matched {n}/{len(manual_names)} manual techs.\n", file=sys.stderr)
    return matched


# ── Comparison logic ────────────────────────────────────────────────────────


@dataclasses.dataclass
class CellResult:
    tech_id: str
    tech_name: str
    manual_name: str
    dimension: str
    pipeline_bin: str | None
    manual_bin: str | None
    r2_bin: str | None
    agree: bool
    r2_agrees_manual: bool
    direction: int          # +1 pipe higher, -1 pipe lower, 0 same
    step_size: int          # abs difference in BIN_ORDER
    evidence_gap: bool
    needs_review: bool
    evidence_strength: float | None
    n_sources: int | None
    single_rater: bool


@dataclasses.dataclass
class ComparisonResult:
    cells: list[CellResult]
    unmatched_manual: list[str]
    unmatched_pipeline: list[str]
    total: int = 0
    n_agree: int = 0
    n_disagree: int = 0
    n_pipeline_lower: int = 0
    n_pipeline_higher: int = 0
    n_r2_agrees_manual: int = 0
    by_dimension: dict[str, list[CellResult]] = dataclasses.field(default_factory=dict)
    by_tech: dict[str, list[CellResult]] = dataclasses.field(default_factory=dict)
    pipeline_file: str = ""
    manual_file: str = ""


def compare(
    pipeline: dict[str, dict],
    manual: dict[str, dict[str, str]],
    matched: dict[str, str],
) -> ComparisonResult:
    """Build cell-by-cell comparisons for all matched tech x dimension pairs."""
    cells = []
    matched_pipeline_ids = set(matched.values())
    unmatched_manual = [m for m in manual if m not in matched]
    unmatched_pipeline = [
        f"{tid} ({d['tech_name']})"
        for tid, d in pipeline.items()
        if tid not in matched_pipeline_ids
    ]

    for mname, tid in sorted(matched.items(), key=lambda kv: kv[1]):
        pdata = pipeline[tid]
        mdata = manual[mname]
        for dim in DIMENSIONS:
            pdim = pdata["dims"].get(dim)
            mbin = mdata.get(dim)
            if pdim is None or mbin is None:
                continue

            pbin = pdim["bin"]
            r2 = pdim["r2_bin"]
            if pbin is None or mbin is None:
                continue

            agree = pbin == mbin
            r2_agrees = (not agree) and (r2 == mbin) if r2 else False
            p_ord = BIN_ORDER.get(pbin, -1)
            m_ord = BIN_ORDER.get(mbin, -1)
            direction = 0
            if p_ord > m_ord:
                direction = 1
            elif p_ord < m_ord:
                direction = -1
            step_size = abs(p_ord - m_ord)

            cells.append(CellResult(
                tech_id=tid,
                tech_name=pdata["tech_name"],
                manual_name=mname,
                dimension=dim,
                pipeline_bin=pbin,
                manual_bin=mbin,
                r2_bin=r2,
                agree=agree,
                r2_agrees_manual=r2_agrees,
                direction=direction,
                step_size=step_size,
                evidence_gap=pdim["evidence_gap"],
                needs_review=pdim["needs_review"],
                evidence_strength=pdim["evidence_strength"],
                n_sources=pdim["n_sources"],
                single_rater=pdim["single_rater"],
            ))

    result = ComparisonResult(
        cells=cells,
        unmatched_manual=unmatched_manual,
        unmatched_pipeline=unmatched_pipeline,
    )
    result.total = len(cells)
    result.n_agree = sum(1 for c in cells if c.agree)
    result.n_disagree = result.total - result.n_agree
    result.n_pipeline_lower = sum(1 for c in cells if c.direction < 0)
    result.n_pipeline_higher = sum(1 for c in cells if c.direction > 0)
    result.n_r2_agrees_manual = sum(1 for c in cells if c.r2_agrees_manual)
    for c in cells:
        result.by_dimension.setdefault(c.dimension, []).append(c)
        result.by_tech.setdefault(c.tech_id, []).append(c)
    return result


# ── Insight generation ──────────────────────────────────────────────────────


def generate_insights(result: ComparisonResult) -> list[str]:
    """Detect patterns and return plain-English insight strings."""
    insights = []
    disagreements = [c for c in result.cells if not c.agree]
    n_dis = len(disagreements)
    if n_dis == 0:
        insights.append("Perfect agreement across all technologies and dimensions.")
        return insights

    # 1. Direction bias
    lower = sum(1 for c in disagreements if c.direction < 0)
    higher = n_dis - lower
    dominant = max(lower, higher)
    if n_dis >= 3 and dominant / n_dis > 0.70:
        label = "lower" if lower > higher else "higher"
        insights.append(
            f"<strong>Systematic {'conservative ' if label == 'lower' else 'optimistic '}"
            f"bias.</strong> Of {n_dis} disagreements, {dominant} have the pipeline "
            f"rating {label} than manual ({dominant * 100 // n_dis}%). "
            f"Only {n_dis - dominant} rate {'higher' if label == 'lower' else 'lower'}."
        )

    # 2. Best and worst dimensions
    dim_rates = {}
    for dim in DIMENSIONS:
        cells_d = result.by_dimension.get(dim, [])
        if cells_d:
            n_a = sum(1 for c in cells_d if c.agree)
            dim_rates[dim] = (n_a, len(cells_d))

    if dim_rates:
        best = max(dim_rates, key=lambda d: dim_rates[d][0] / dim_rates[d][1])
        worst = min(dim_rates, key=lambda d: dim_rates[d][0] / dim_rates[d][1])
        ba, bt = dim_rates[best]
        wa, wt = dim_rates[worst]

        if ba == bt:
            insights.append(
                f"<strong>{best} is solid.</strong> Perfect {ba}/{bt} agreement."
            )
        elif ba / bt >= 0.8:
            insights.append(
                f"<strong>{best} has the strongest agreement</strong> at {ba}/{bt} "
                f"({ba * 100 // bt}%)."
            )

        if wa == 0:
            insights.append(
                f"<strong>{worst} is a total miss.</strong> Zero agreement across "
                f"all {wt} techs. This points to a systematic prompt or rubric issue, "
                f"not per-tech evidence problems."
            )
        elif wa / wt <= 0.5 and worst != best:
            insights.append(
                f"<strong>{worst} has the lowest agreement</strong> at {wa}/{wt} "
                f"({wa * 100 // wt}%)."
            )

    # 3. R2 correction pattern
    r2_corrections = [c for c in disagreements if c.r2_agrees_manual]
    if n_dis >= 3 and len(r2_corrections) / n_dis > 0.30:
        dim_counts = Counter(c.dimension for c in r2_corrections)
        top_dims = dim_counts.most_common(2)
        dim_note = ", ".join(
            f"{d} ({n} of {len(result.by_dimension.get(d, []))})"
            for d, n in top_dims
        )
        insights.append(
            f"<strong>R2 (second rater) corrects R1 toward manual in "
            f"{len(r2_corrections)} of {n_dis} cases.</strong> "
            f"This is especially visible on {dim_note}. "
            f"The merge logic defaults to R1, so these corrections only surface "
            f"when raters disagree."
        )

    # 4. Evidence gap clustering
    gap_dis = sum(1 for c in disagreements if c.evidence_gap)
    if gap_dis > 0:
        pct = gap_dis * 100 // n_dis
        qualifier = "most" if pct >= 60 else "many of the" if pct >= 40 else "some"
        insights.append(
            f"<strong>Evidence gaps cluster with disagreements.</strong> "
            f"{gap_dis} of {n_dis} disagreements ({pct}%) have an evidence gap "
            f"flagged. The pipeline's ratings may partly reflect thin evidence "
            f"on dimensions that rely on industry or trade-press sources."
        )

    # 5. Outlier detection (2-step gaps)
    outliers = [c for c in disagreements if c.step_size >= 2]
    for c in outliers:
        dir_label = "more optimistic" if c.direction > 0 else "more conservative"
        insights.append(
            f"<strong>Large disagreement: {c.tech_name} {c.dimension}.</strong> "
            f"Pipeline={c.pipeline_bin}, manual={c.manual_bin} (2-step gap). "
            f"R2={c.r2_bin or 'n/a'}. The pipeline is {dir_label} here — "
            f"worth investigating whether evidence was conflated across technologies."
        )

    # 6. Per-tech extremes
    for tid, cells_t in result.by_tech.items():
        if all(c.agree for c in cells_t) and len(cells_t) == len(DIMENSIONS):
            insights.append(
                f"<strong>{cells_t[0].tech_name}:</strong> perfect agreement "
                f"across all {len(DIMENSIONS)} dimensions."
            )
        elif not any(c.agree for c in cells_t) and len(cells_t) >= 4:
            insights.append(
                f"<strong>{cells_t[0].tech_name}:</strong> disagrees on every "
                f"dimension ({len(cells_t)}/{len(cells_t)})."
            )

    return insights


# ── CSS ─────────────────────────────────────────────────────────────────────

CSS = """\
:root {
  --bg: #F7F5F0;
  --bg-card: #EFECE6;
  --bg-cell: #E8E5DE;
  --text-primary: #3A3632;
  --text-secondary: #6B665E;
  --text-muted: #9B9589;
  --border: #D6D1C8;
  --agree: #6B8F71;
  --agree-bg: #E8F0E9;
  --disagree: #C4644A;
  --disagree-bg: #F5E8E3;
  --r2-manual: #5B7FA5;
  --r2-manual-bg: #E4ECF3;
  --no-data: #C8C3B9;
  --no-data-bg: #EDEBE5;
  --heading-font: Georgia, 'Times New Roman', serif;
  --body-font: -apple-system, BlinkMacSystemFont, 'Segoe UI', system-ui, sans-serif;
  --mono-font: 'SF Mono', 'Cascadia Code', 'Fira Code', Menlo, monospace;
}
@media (prefers-color-scheme: dark) {
  :root {
    --bg: #1E1F23; --bg-card: #282A2E; --bg-cell: #32343A;
    --text-primary: #D4D0C8; --text-secondary: #9B9790; --text-muted: #6E6A63;
    --border: #3E4046;
    --agree: #7DA882; --agree-bg: #2A3A2C;
    --disagree: #D47A62; --disagree-bg: #3A2A24;
    --r2-manual: #7BA0C4; --r2-manual-bg: #24303E;
    --no-data: #555248; --no-data-bg: #2A2B2F;
  }
}
:root[data-theme="dark"] {
  --bg: #1E1F23; --bg-card: #282A2E; --bg-cell: #32343A;
  --text-primary: #D4D0C8; --text-secondary: #9B9790; --text-muted: #6E6A63;
  --border: #3E4046;
  --agree: #7DA882; --agree-bg: #2A3A2C;
  --disagree: #D47A62; --disagree-bg: #3A2A24;
  --r2-manual: #7BA0C4; --r2-manual-bg: #24303E;
  --no-data: #555248; --no-data-bg: #2A2B2F;
}
:root[data-theme="light"] {
  --bg: #F7F5F0; --bg-card: #EFECE6; --bg-cell: #E8E5DE;
  --text-primary: #3A3632; --text-secondary: #6B665E; --text-muted: #9B9589;
  --border: #D6D1C8;
  --agree: #6B8F71; --agree-bg: #E8F0E9;
  --disagree: #C4644A; --disagree-bg: #F5E8E3;
  --r2-manual: #5B7FA5; --r2-manual-bg: #E4ECF3;
  --no-data: #C8C3B9; --no-data-bg: #EDEBE5;
}
* { margin: 0; padding: 0; box-sizing: border-box; }
body {
  background: var(--bg); color: var(--text-primary);
  font-family: var(--body-font); font-size: 15px; line-height: 1.55;
  padding: 2rem 1.5rem 4rem; max-width: 960px; margin: 0 auto;
}
h1 {
  font-family: var(--heading-font); font-size: 1.6rem; font-weight: 700;
  letter-spacing: -0.02em; margin-bottom: 0.25rem;
}
.subtitle { color: var(--text-secondary); font-size: 0.85rem; margin-bottom: 2rem; }
.stats {
  display: grid; grid-template-columns: repeat(auto-fit, minmax(140px, 1fr));
  gap: 10px; margin-bottom: 2.5rem;
}
.stat {
  background: var(--bg-card); border: 1px solid var(--border);
  border-radius: 6px; padding: 14px 16px;
}
.stat-value {
  font-family: var(--mono-font); font-size: 1.7rem; font-weight: 700;
  font-variant-numeric: tabular-nums; line-height: 1.1;
}
.stat-value.agree-color { color: var(--agree); }
.stat-value.disagree-color { color: var(--disagree); }
.stat-value.r2-color { color: var(--r2-manual); }
.stat-label {
  font-size: 0.72rem; text-transform: uppercase; letter-spacing: 0.06em;
  color: var(--text-muted); margin-top: 4px;
}
h2 {
  font-family: var(--heading-font); font-size: 1.15rem; font-weight: 700;
  margin-bottom: 0.15rem;
}
.section-desc { color: var(--text-secondary); font-size: 0.82rem; margin-bottom: 1rem; }
.heatmap-wrap {
  overflow-x: auto; margin-bottom: 2.5rem; -webkit-overflow-scrolling: touch;
}
.heatmap { border-collapse: separate; border-spacing: 3px; width: 100%; min-width: 620px; }
.heatmap th {
  font-family: var(--mono-font); font-size: 0.72rem; font-weight: 600;
  text-transform: uppercase; letter-spacing: 0.06em; color: var(--text-muted);
  padding: 6px 8px; text-align: center; white-space: nowrap;
}
.heatmap th.tech-header {
  text-align: left; font-family: var(--body-font); text-transform: none;
  letter-spacing: 0; font-size: 0.8rem; color: var(--text-secondary);
  max-width: 220px; overflow: hidden; text-overflow: ellipsis;
}
.heatmap td {
  text-align: center; padding: 7px 6px; border-radius: 4px;
  font-family: var(--mono-font); font-size: 0.78rem; font-weight: 600;
  font-variant-numeric: tabular-nums; position: relative; cursor: default;
  min-width: 72px;
}
.heatmap td.agree { background: var(--agree-bg); color: var(--agree); }
.heatmap td.disagree { background: var(--disagree-bg); color: var(--disagree); }
.heatmap td.r2-manual { background: var(--r2-manual-bg); color: var(--r2-manual); }
.heatmap td.no-data { background: var(--no-data-bg); color: var(--no-data); font-style: italic; font-weight: 400; }
.heatmap td .cell-sub { display: block; font-size: 0.62rem; font-weight: 400; opacity: 0.75; margin-top: 1px; }
.heatmap td[data-tip] { position: relative; }
.heatmap td[data-tip]:hover::after {
  content: attr(data-tip); position: absolute; bottom: calc(100% + 6px);
  left: 50%; transform: translateX(-50%); background: var(--text-primary);
  color: var(--bg); font-family: var(--body-font); font-size: 0.7rem;
  font-weight: 400; padding: 5px 9px; border-radius: 4px; z-index: 10;
  pointer-events: none; line-height: 1.35; max-width: 280px; white-space: normal;
}
.legend {
  display: flex; flex-wrap: wrap; gap: 16px; margin-bottom: 1.5rem;
  font-size: 0.78rem; color: var(--text-secondary);
}
.legend-item { display: flex; align-items: center; gap: 6px; }
.legend-swatch { width: 14px; height: 14px; border-radius: 3px; flex-shrink: 0; }
.dim-bar-section { margin-bottom: 2.5rem; }
.dim-bars { display: flex; flex-direction: column; gap: 8px; }
.dim-row { display: flex; align-items: center; gap: 10px; }
.dim-label {
  font-family: var(--mono-font); font-size: 0.78rem; font-weight: 600;
  width: 36px; flex-shrink: 0; color: var(--text-secondary);
}
.dim-bar-track {
  flex: 1; height: 22px; background: var(--bg-cell);
  border-radius: 4px; display: flex; overflow: hidden; gap: 2px;
}
.dim-bar-seg { height: 100%; }
.dim-bar-seg.seg-agree { background: var(--agree); }
.dim-bar-seg.seg-disagree { background: var(--disagree); }
.dim-bar-seg.seg-r2 { background: var(--r2-manual); }
.dim-bar-seg.seg-nodata { background: var(--no-data); opacity: 0.5; }
.dim-count {
  font-family: var(--mono-font); font-size: 0.72rem; color: var(--text-muted);
  width: 70px; flex-shrink: 0; text-align: right; font-variant-numeric: tabular-nums;
}
.detail-wrap { overflow-x: auto; margin-bottom: 2.5rem; -webkit-overflow-scrolling: touch; }
.detail-table { width: 100%; min-width: 680px; border-collapse: collapse; font-size: 0.8rem; }
.detail-table th {
  font-size: 0.7rem; text-transform: uppercase; letter-spacing: 0.05em;
  color: var(--text-muted); font-weight: 600; text-align: left;
  padding: 8px 10px; border-bottom: 2px solid var(--border); white-space: nowrap;
}
.detail-table td {
  padding: 7px 10px; border-bottom: 1px solid var(--border);
  font-variant-numeric: tabular-nums; vertical-align: top;
}
.detail-table tr:last-child td { border-bottom: none; }
.detail-table .mono { font-family: var(--mono-font); font-size: 0.78rem; }
.bin-badge {
  display: inline-block; padding: 1px 7px; border-radius: 3px;
  font-family: var(--mono-font); font-size: 0.72rem; font-weight: 600;
}
.bin-badge.high { background: var(--agree-bg); color: var(--agree); }
.bin-badge.mid { background: var(--r2-manual-bg); color: var(--r2-manual); }
.bin-badge.low { background: var(--disagree-bg); color: var(--disagree); }
.direction { font-size: 0.72rem; color: var(--text-muted); font-style: italic; }
.tag {
  display: inline-block; padding: 1px 6px; border-radius: 3px;
  font-size: 0.68rem; font-weight: 600; text-transform: uppercase;
  letter-spacing: 0.04em; margin-left: 4px;
}
.tag-evgap { background: #F5E0D0; color: #A05A2C; }
.tag-review { background: #F5E8E3; color: var(--disagree); }
@media (prefers-color-scheme: dark) {
  .tag-evgap { background: #3A2E24; color: #D4A06C; }
  .tag-review { background: #3A2A24; color: var(--disagree); }
}
:root[data-theme="dark"] .tag-evgap { background: #3A2E24; color: #D4A06C; }
:root[data-theme="dark"] .tag-review { background: #3A2A24; color: var(--disagree); }
:root[data-theme="light"] .tag-evgap { background: #F5E0D0; color: #A05A2C; }
:root[data-theme="light"] .tag-review { background: #F5E8E3; color: var(--disagree); }
.patterns {
  background: var(--bg-card); border: 1px solid var(--border);
  border-radius: 6px; padding: 20px 22px; margin-bottom: 2.5rem;
}
.patterns h3 { font-family: var(--heading-font); font-size: 0.95rem; margin-bottom: 12px; }
.pattern-item {
  margin-bottom: 10px; padding-left: 14px;
  border-left: 3px solid var(--border); font-size: 0.85rem; line-height: 1.5;
}
.pattern-item:last-child { margin-bottom: 0; }
.footnote {
  font-size: 0.75rem; color: var(--text-muted);
  border-top: 1px solid var(--border); padding-top: 12px; line-height: 1.5;
}
@media (max-width: 600px) {
  body { padding: 1.2rem 1rem 3rem; }
  h1 { font-size: 1.3rem; }
  .stats { grid-template-columns: repeat(2, 1fr); }
}
"""


# ── HTML rendering ──────────────────────────────────────────────────────────


def _h(text: str) -> str:
    """HTML-escape a string."""
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _bin_badge(b: str | None) -> str:
    if not b:
        return "-"
    cls = b.lower() if b in ("High", "Mid", "Low") else ""
    return f'<span class="bin-badge {cls}">{_h(b)}</span>'


def _build_stat_tiles(r: ComparisonResult) -> str:
    pct_agree = r.n_agree * 100 // r.total if r.total else 0
    pct_dis = 100 - pct_agree
    tiles = [
        ("", str(r.total), "Comparisons"),
        ("agree-color", f"{pct_agree}%", "Agree with manual"),
        ("disagree-color", f"{pct_dis}%", "Disagree"),
        ("disagree-color", str(r.n_pipeline_lower), "Pipeline lower"),
        ("r2-color", str(r.n_r2_agrees_manual), "R2 agrees w/ manual"),
    ]
    parts = []
    for cls, val, label in tiles:
        parts.append(
            f'<div class="stat">'
            f'<div class="stat-value {cls}">{val}</div>'
            f'<div class="stat-label">{label}</div>'
            f'</div>'
        )
    return "\n".join(parts)


def _build_heatmap(r: ComparisonResult) -> str:
    tech_order = []
    seen = set()
    for c in r.cells:
        if c.tech_id not in seen:
            seen.add(c.tech_id)
            tech_order.append((c.tech_id, c.tech_name))

    cell_map = {(c.tech_id, c.dimension): c for c in r.cells}

    rows = []
    for tid, tname in tech_order:
        cols = []
        for dim in DIMENSIONS:
            c = cell_map.get((tid, dim))
            if c is None:
                cols.append(
                    f'<td class="no-data" data-tip="No pipeline data">n/a</td>'
                )
                continue

            if c.agree:
                cls = "agree"
                tip = f"Manual: {c.manual_bin}, R1: {c.pipeline_bin}, R2: {c.r2_bin or 'n/a'}"
                inner = _h(c.pipeline_bin)
            elif c.r2_agrees_manual:
                cls = "r2-manual"
                tip = (
                    f"Manual: {c.manual_bin}, R1: {c.pipeline_bin}, "
                    f"R2: {c.r2_bin} — R2 sides with manual"
                )
                inner = (
                    f'{_h(c.pipeline_bin)}'
                    f'<span class="cell-sub">man: {_h(c.manual_bin)}</span>'
                )
            else:
                cls = "disagree"
                r2_note = f", R2: {c.r2_bin}" if c.r2_bin else ""
                tip = (
                    f"Manual: {c.manual_bin}, R1: {c.pipeline_bin}"
                    f"{r2_note} — both raters differ"
                )
                inner = (
                    f'{_h(c.pipeline_bin)}'
                    f'<span class="cell-sub">man: {_h(c.manual_bin)}</span>'
                )

            if c.evidence_gap:
                tip += " [evidence gap]"

            cols.append(f'<td class="{cls}" data-tip="{_h(tip)}">{inner}</td>')

        label = _h(tname)
        if len(tname) > 30:
            label = _h(tname[:28]) + "..."
        rows.append(
            f'<tr><th class="tech-header">{label}</th>{"".join(cols)}</tr>'
        )

    header = (
        '<tr><th class="tech-header">Technology</th>'
        + "".join(f"<th>{d}</th>" for d in DIMENSIONS)
        + "</tr>"
    )
    return (
        f'<table class="heatmap"><thead>{header}</thead>'
        f'<tbody>{"".join(rows)}</tbody></table>'
    )


def _build_dimension_bars(r: ComparisonResult) -> str:
    parts = []
    for dim in DIMENSIONS:
        cells_d = r.by_dimension.get(dim, [])
        total = len(cells_d)
        if total == 0:
            continue
        n_agree = sum(1 for c in cells_d if c.agree)
        n_r2 = sum(1 for c in cells_d if c.r2_agrees_manual)
        n_dis = total - n_agree - n_r2

        def pct(n):
            return f"{n * 100 / total:.1f}%" if n > 0 else "0%"

        segs = ""
        if n_agree:
            segs += f'<div class="dim-bar-seg seg-agree" style="width:{pct(n_agree)}"></div>'
        if n_r2:
            segs += f'<div class="dim-bar-seg seg-r2" style="width:{pct(n_r2)}"></div>'
        if n_dis:
            segs += f'<div class="dim-bar-seg seg-disagree" style="width:{pct(n_dis)}"></div>'

        parts.append(
            f'<div class="dim-row">'
            f'<span class="dim-label">{dim}</span>'
            f'<div class="dim-bar-track">{segs}</div>'
            f'<span class="dim-count">{n_agree}/{total} agree</span>'
            f'</div>'
        )
    return "\n".join(parts)


def _build_insights_html(insights: list[str]) -> str:
    if not insights:
        return '<div class="pattern-item">No notable patterns detected.</div>'
    return "\n".join(f'<div class="pattern-item">{s}</div>' for s in insights)


def _build_detail_table(r: ComparisonResult) -> str:
    disagreements = sorted(
        [c for c in r.cells if not c.agree],
        key=lambda c: (DIMENSIONS.index(c.dimension), c.tech_id),
    )
    if not disagreements:
        return "<p>No disagreements.</p>"

    header = (
        "<thead><tr>"
        "<th>Technology</th><th>Dim</th><th>Manual</th>"
        "<th>R1</th><th>R2</th><th>Direction</th><th>Flags</th>"
        "</tr></thead>"
    )
    rows = []
    for c in disagreements:
        if c.direction < 0:
            if c.r2_agrees_manual:
                direction = "pipe lower, R2 = manual"
            else:
                direction = "pipe lower, both raters"
        elif c.direction > 0:
            if c.r2_agrees_manual:
                direction = "pipe higher, R2 = manual"
            else:
                direction = "pipe higher, both raters"
        else:
            direction = ""
        if c.step_size >= 2:
            direction = f"pipe {'+' if c.direction > 0 else ''}{c.step_size} steps"

        flags = ""
        if c.evidence_gap:
            flags += '<span class="tag tag-evgap">ev gap</span>'
        if c.needs_review:
            flags += '<span class="tag tag-review">review</span>'

        rows.append(
            f"<tr>"
            f"<td>{_h(c.tech_name)}</td>"
            f'<td class="mono">{c.dimension}</td>'
            f"<td>{_bin_badge(c.manual_bin)}</td>"
            f"<td>{_bin_badge(c.pipeline_bin)}</td>"
            f"<td>{_bin_badge(c.r2_bin)}</td>"
            f'<td class="direction">{direction}</td>'
            f"<td>{flags}</td>"
            f"</tr>"
        )
    return f'{header}<tbody>{"".join(rows)}</tbody>'


def _build_footer(r: ComparisonResult) -> str:
    parts = []
    n_matched = len(r.by_tech)
    parts.append(f"{n_matched} technologies matched between pipeline and manual.")
    if r.unmatched_manual:
        names = ", ".join(f'"{n}"' for n in r.unmatched_manual)
        parts.append(f" Manual-only (not in pipeline): {names}.")
    if r.unmatched_pipeline:
        names = ", ".join(r.unmatched_pipeline)
        parts.append(f" Pipeline-only (not in manual): {names}.")
    parts.append(
        ' Pipeline R1 = primary rater, R2 = second rater.'
        ' "Evidence gap" means the rater flagged insufficient direct evidence.'
        ' "Needs review" is the pipeline\'s own flag for human attention.'
    )
    return "".join(parts)


HTML_TEMPLATE = Template("""\
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>BRLa: $title</title>
<style>
$css
</style>
</head>
<body>

<h1>$heading</h1>
<p class="subtitle">$subtitle</p>

<div class="stats">
$stat_tiles
</div>

<h2>Agreement Heatmap</h2>
<p class="section-desc">Each cell shows the pipeline's R1 bin. Color encodes agreement pattern with the manual assessment.</p>

<div class="legend">
  <div class="legend-item"><div class="legend-swatch" style="background:var(--agree)"></div> Both raters agree with manual</div>
  <div class="legend-item"><div class="legend-swatch" style="background:var(--r2-manual)"></div> R2 agrees with manual, R1 differs</div>
  <div class="legend-item"><div class="legend-swatch" style="background:var(--disagree)"></div> Both raters differ from manual</div>
  <div class="legend-item"><div class="legend-swatch" style="background:var(--no-data);opacity:0.5"></div> Not in pipeline</div>
</div>

<div class="heatmap-wrap">
$heatmap
</div>

<div class="dim-bar-section">
<h2>Disagreement Rate by Dimension</h2>
<p class="section-desc">$dim_desc</p>
<div class="dim-bars">
$dimension_bars
</div>
</div>

<div class="patterns">
<h3>Key Patterns</h3>
$insights_html
</div>

<h2>All $n_disagree Disagreements</h2>
<p class="section-desc">Sorted by dimension, then technology.</p>

<div class="detail-wrap">
<table class="detail-table">
$detail_table
</table>
</div>

<p class="footnote">
<strong>Coverage:</strong> $footer
</p>

</body>
</html>
""")


def render_html(result: ComparisonResult, insights: list[str]) -> str:
    pipeline_name = Path(result.pipeline_file).name
    manual_name = Path(result.manual_file).name

    dim_rates = {}
    for dim in DIMENSIONS:
        cells_d = result.by_dimension.get(dim, [])
        if cells_d:
            n_a = sum(1 for c in cells_d if c.agree)
            dim_rates[dim] = (n_a, len(cells_d))

    if dim_rates:
        best = max(dim_rates, key=lambda d: dim_rates[d][0] / dim_rates[d][1])
        worst = min(dim_rates, key=lambda d: dim_rates[d][0] / dim_rates[d][1])
        ba, bt = dim_rates[best]
        wa, wt = dim_rates[worst]
        dim_desc = (
            f"{best} has the strongest agreement ({ba}/{bt}). "
            f"{worst} has the weakest ({wa}/{wt})."
        )
    else:
        dim_desc = ""

    return HTML_TEMPLATE.substitute(
        title=f"Pipeline vs {manual_name}",
        heading="Pipeline vs Manual Assessment",
        subtitle=f"Comparing {pipeline_name} against {manual_name}",
        css=CSS,
        stat_tiles=_build_stat_tiles(result),
        heatmap=_build_heatmap(result),
        dim_desc=dim_desc,
        dimension_bars=_build_dimension_bars(result),
        insights_html=_build_insights_html(insights),
        n_disagree=result.n_disagree,
        detail_table=_build_detail_table(result),
        footer=_build_footer(result),
    )


# ── CLI entry point ─────────────────────────────────────────────────────────


def main():
    ap = argparse.ArgumentParser(
        description="Compare pipeline vs manual BRLa assessments → HTML dashboard."
    )
    ap.add_argument(
        "--pipeline",
        type=Path,
        default=ROOT / "output" / "review.xlsx",
        help="Pipeline review spreadsheet (default: output/review.xlsx)",
    )
    ap.add_argument(
        "--manual",
        type=Path,
        required=True,
        help="Manually-performed BRLa assessment xlsx",
    )
    ap.add_argument(
        "--mapping",
        type=Path,
        default=None,
        help="Optional JSON mapping {manual_name: tech_id}",
    )
    ap.add_argument(
        "--threshold",
        type=float,
        default=0.4,
        help="Fuzzy match threshold (default: 0.4)",
    )
    ap.add_argument(
        "-o", "--output",
        type=Path,
        default=ROOT / "output" / "comparison.html",
        help="Output HTML path (default: output/comparison.html)",
    )
    args = ap.parse_args()

    pipeline = load_pipeline(args.pipeline)
    manual = load_manual(args.manual)

    matched = match_techs(
        list(manual.keys()), pipeline, args.mapping, args.threshold
    )
    if not matched:
        print("Error: no techs matched. Check names or provide --mapping.", file=sys.stderr)
        sys.exit(1)

    result = compare(pipeline, manual, matched)
    result.pipeline_file = str(args.pipeline)
    result.manual_file = str(args.manual)

    insights = generate_insights(result)
    html = render_html(result, insights)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(html, encoding="utf-8")
    print(f"Dashboard written to {args.output}")
    print(f"  {result.total} comparisons: {result.n_agree} agree, {result.n_disagree} disagree")


if __name__ == "__main__":
    main()
