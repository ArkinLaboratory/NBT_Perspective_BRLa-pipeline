"""Foldback — read reviewed review.xlsx, write human fields into master.json.

The reviewer fills in three columns per row in review.xlsx:
  human_bin    (Low / Mid / High or blank)
  human_notes  (free text or blank)
  reviewed     (any truthy value: "yes", "x", "1", TRUE)

This script reads those columns, matches rows by (tech_id, dimension),
and writes the human fields back into master.json.  The original master.json
is backed up before overwriting.
"""
import shutil
from pathlib import Path

from openpyxl import load_workbook

from .utils import read_json, write_json

VALID_BINS = {"Low", "Mid", "High"}


def _is_truthy(val) -> bool:
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    return str(val).strip().lower() in ("yes", "y", "x", "1", "true")


def read_reviewed_xlsx(xlsx_path: Path) -> dict[tuple[str, str], dict]:
    """Return {(tech_id, dimension): {human_bin, human_notes, reviewed}}."""
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}

    required = {"tech_id", "dimension", "human_bin", "human_notes", "reviewed"}
    missing = required - set(col.keys())
    if missing:
        raise RuntimeError(
            f"review.xlsx missing columns: {missing}. "
            f"Found: {list(col.keys())}")

    results = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        tech_id = row[col["tech_id"]]
        dim = row[col["dimension"]]
        if not tech_id or not dim:
            continue

        human_bin = row[col["human_bin"]]
        if human_bin and str(human_bin).strip() not in VALID_BINS:
            human_bin = None

        human_notes = row[col["human_notes"]]
        if human_notes:
            human_notes = str(human_notes).strip()

        reviewed = _is_truthy(row[col["reviewed"]])

        if human_bin or human_notes or reviewed:
            results[(str(tech_id).strip(), str(dim).strip())] = {
                "human_bin": str(human_bin).strip() if human_bin else None,
                "human_notes": human_notes or None,
                "reviewed": reviewed,
            }

    wb.close()
    return results


def run(cfg: dict) -> dict:
    out_dir = cfg["paths"]["output_dir"]
    master_path = out_dir / "master.json"
    xlsx_path = out_dir / "review.xlsx"

    if not master_path.exists():
        raise RuntimeError(f"master.json not found at {master_path}")
    if not xlsx_path.exists():
        raise RuntimeError(f"review.xlsx not found at {xlsx_path}")

    backup_path = master_path.with_suffix(".json.bak")
    shutil.copy2(master_path, backup_path)

    master = read_json(master_path)
    reviewed = read_reviewed_xlsx(xlsx_path)

    updated = 0
    for entry in master["entries"]:
        key = (entry["tech_id"], entry["dimension"])
        if key in reviewed:
            human = reviewed[key]
            entry["human_bin"] = human["human_bin"]
            entry["human_notes"] = human["human_notes"]
            entry["reviewed"] = human["reviewed"]
            updated += 1

    write_json(master_path, master)
    return {
        "n_reviewed_rows": len(reviewed),
        "n_updated": updated,
        "backup": str(backup_path),
    }
