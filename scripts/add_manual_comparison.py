#!/usr/bin/env python
"""Add a manual_compare column to an existing review.xlsx.

Reads a manual-assessment xlsx (with a tech_id column) and the pipeline's
review.xlsx, computes agreement between each LLM rater and the manual
assignment, and writes a new xlsx with the column inserted after needs_review.

Rebuilds the workbook from scratch to preserve formatting. The original
review.xlsx is not modified when --output is specified.

Usage:
    conda run -n env-brla python scripts/add_manual_comparison.py \
        --manual notes/assets/sheets/brla_manual.xlsx \
        --review output/review.xlsx \
        --output output/review_with-manual-compare.xlsx

    python scripts/add_manual_comparison.py \
        --manual notes/assets/sheets/brla_manual.xlsx \
        --review output/review.xlsx \
        --output output/review_with-manual-compare.xlsx
"""
import argparse
import shutil
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DIMENSIONS = ["TRL", "MRL", "RRL", "ARL", "ORL"]

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4",
                          fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
REVIEW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC",
                          fill_type="solid")
LINK_FONT = Font(color="0563C1", underline="single", size=9)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
TOP_ALIGN = Alignment(vertical="top")

MC_ALL_AGREE = PatternFill(start_color="D5F5D5", end_color="D5F5D5",
                           fill_type="solid")
MC_R1_DISAGREE = PatternFill(start_color="FFF2CC", end_color="FFF2CC",
                             fill_type="solid")
MC_R2_DISAGREE = PatternFill(start_color="D6EAF8", end_color="D6EAF8",
                             fill_type="solid")
MC_MANUAL_DISAGREE = PatternFill(start_color="F5D5D5", end_color="F5D5D5",
                                 fill_type="solid")
MC_ALL_DISAGREE = PatternFill(start_color="E8E8E8", end_color="E8E8E8",
                              fill_type="solid")

WRAP_COLS = {"rater_1_rationale", "rater_2_rationale", "manual_compare"}
MC_COL = "manual_compare"
INSERT_AFTER = "needs_review"


def load_manual(path: Path) -> dict[tuple[str, str], str]:
    """Read manual draft xlsx into {(tech_id, dimension): bin}."""
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}

    if "tech_id" not in col:
        raise RuntimeError(
            f"Manual draft missing 'tech_id' column. Found: {list(col.keys())}")

    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        tid = row[col["tech_id"]]
        if not tid:
            continue
        tid = str(tid).strip()
        for dim in DIMENSIONS:
            if dim in col:
                val = row[col[dim]]
                if val:
                    result[(tid, dim)] = str(val).strip()
    wb.close()
    return result


def compare_cell(r1_bin: str, r2_bin: str,
                 manual_bin: str) -> tuple[str, PatternFill | None]:
    """Return (cell_text, fill) for a manual comparison cell."""
    r1 = (r1_bin or "").strip() or None
    r2 = (r2_bin or "").strip() or None
    m = (manual_bin or "").strip() or None

    if not m:
        return "", None

    if r1 == r2 == m:
        return f"ALL AGREE = {m}", MC_ALL_AGREE
    if r1 == r2 and r1 != m:
        return f"Rater_1 = Rater_2 = {r1}\nManual = {m}", MC_MANUAL_DISAGREE
    if r2 == m and r1 != m:
        return f"Rater_2 = Manual = {m}\nRater_1 = {r1}", MC_R1_DISAGREE
    if r1 == m and r2 != m:
        return f"Rater_1 = Manual = {m}\nRater_2 = {r2}", MC_R2_DISAGREE
    return f"Rater_1 = {r1}\nRater_2 = {r2}\nManual = {m}", MC_ALL_DISAGREE


def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--manual", required=True,
                   help="path to manual-assessment xlsx (must have tech_id column)")
    p.add_argument("--review", default="output/review.xlsx",
                   help="path to pipeline review.xlsx (default: output/review.xlsx)")
    p.add_argument("--output", default=None,
                   help="write to a different file instead of modifying review "
                        "in place (default: modify --review)")
    args = p.parse_args()

    manual_path = Path(args.manual)
    review_path = Path(args.review)
    output_path = Path(args.output) if args.output else review_path

    if not manual_path.exists():
        raise FileNotFoundError(f"Manual draft not found: {manual_path}")
    if not review_path.exists():
        raise FileNotFoundError(f"review.xlsx not found: {review_path}")

    manual = load_manual(manual_path)
    print(f"Manual draft: {len(manual)} entries from {manual_path}")

    if output_path == review_path:
        shutil.copy2(review_path, review_path.with_suffix(".xlsx.bak"))

    # -- Read source workbook (data + structure) --
    src_wb = load_workbook(review_path)
    src_ws = src_wb.active
    src_headers = [cell.value for cell in src_ws[1]]
    src_col = {h: i for i, h in enumerate(src_headers)}

    for req in ("tech_id", "dimension", "rater_1_bin", "rater_2_bin", INSERT_AFTER):
        if req not in src_col:
            raise RuntimeError(f"review.xlsx missing column: {req}")

    # Strip existing manual_compare if re-running
    if MC_COL in src_headers:
        src_headers = [h for h in src_headers if h != MC_COL]
        src_col = {h: i for i, h in enumerate(src_headers)}

    # Build new header list with manual_compare inserted after needs_review
    insert_pos = src_headers.index(INSERT_AFTER) + 1
    new_headers = src_headers[:insert_pos] + [MC_COL] + src_headers[insert_pos:]

    # Read source column widths
    src_widths = {}
    for col_idx in range(1, len(src_headers) + 1):
        letter = get_column_letter(col_idx)
        dim = src_ws.column_dimensions.get(letter)
        if dim and dim.width:
            src_widths[src_headers[col_idx - 1]] = dim.width

    source_names = [h for h in src_headers if h and h.startswith("source_")]
    if source_names and source_names[0] in src_widths:
        for name in source_names[1:]:
            src_widths.setdefault(name, src_widths[source_names[0]])

    # Read source rows: values + hyperlinks + per-row needs_review flag
    src_rows = []
    src_hyperlinks = []
    for row_idx in range(2, src_ws.max_row + 1):
        row_vals = {}
        row_links = {}
        for col_name in src_headers:
            orig_col_idx = src_col[col_name]
            cell = src_ws.cell(row=row_idx, column=orig_col_idx + 1)
            row_vals[col_name] = cell.value
            if cell.hyperlink:
                row_links[col_name] = cell.hyperlink.target
        src_rows.append(row_vals)
        src_hyperlinks.append(row_links)
    src_wb.close()

    # -- Compute manual comparison for each row --
    counts = {"all_agree": 0, "r1_disagree": 0, "r2_disagree": 0,
              "manual_disagree": 0, "all_disagree": 0, "no_manual": 0}
    mc_data = []
    for row_vals in src_rows:
        tid = str(row_vals.get("tech_id", "") or "").strip()
        dim = str(row_vals.get("dimension", "") or "").strip()
        r1 = str(row_vals.get("rater_1_bin", "") or "").strip()
        r2 = str(row_vals.get("rater_2_bin", "") or "").strip()
        manual_bin = manual.get((tid, dim), "")
        text, fill = compare_cell(r1, r2, manual_bin)
        mc_data.append((text, fill))

        if not manual_bin:
            counts["no_manual"] += 1
        elif fill == MC_ALL_AGREE:
            counts["all_agree"] += 1
        elif fill == MC_R1_DISAGREE:
            counts["r1_disagree"] += 1
        elif fill == MC_R2_DISAGREE:
            counts["r2_disagree"] += 1
        elif fill == MC_MANUAL_DISAGREE:
            counts["manual_disagree"] += 1
        elif fill == MC_ALL_DISAGREE:
            counts["all_disagree"] += 1

    # -- Write new workbook --
    wb = Workbook()
    ws = wb.active
    ws.title = "BRLa Review"
    ws.freeze_panes = "A2"

    # Headers
    for col_idx, name in enumerate(new_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = TOP_ALIGN

    # Data rows
    for row_num, (row_vals, row_links, (mc_text, mc_fill)) in enumerate(
            zip(src_rows, src_hyperlinks, mc_data), 2):

        needs_review = str(row_vals.get("needs_review", "") or "").strip() == "Yes"

        for col_idx, col_name in enumerate(new_headers, 1):
            if col_name == MC_COL:
                val = mc_text
            else:
                val = row_vals.get(col_name, "")

            cell = ws.cell(row=row_num, column=col_idx, value=val)

            # Alignment
            if col_name in WRAP_COLS or col_name.startswith("source_"):
                cell.alignment = WRAP_ALIGN
            else:
                cell.alignment = TOP_ALIGN

            # Hyperlinks on source columns
            if col_name in row_links:
                cell.hyperlink = row_links[col_name]
                cell.font = LINK_FONT

            # needs_review yellow highlight
            if needs_review and col_name != MC_COL:
                cell.fill = REVIEW_FILL

            # Manual compare fill (overrides review fill for this cell)
            if col_name == MC_COL and mc_fill:
                cell.fill = mc_fill

    # Group source columns — then set all widths (group() wipes ColumnDimension objects)
    source_indices = [i for i, h in enumerate(new_headers, 1)
                      if h and h.startswith("source_")]
    if source_indices:
        ws.column_dimensions.group(
            get_column_letter(source_indices[0]),
            get_column_letter(source_indices[-1]),
            hidden=False)

    for col_idx, name in enumerate(new_headers, 1):
        width = 28 if name == MC_COL else src_widths.get(name, 12)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(new_headers))}{len(src_rows) + 1}")

    wb.save(output_path)
    print(f"Wrote {output_path}")
    if output_path != review_path:
        print(f"  Source: {args.review}")
    else:
        print(f"  Backup: {output_path.with_suffix('.xlsx.bak')}")
    print(f"\n  Results ({sum(v for k, v in counts.items() if k != 'no_manual')}"
          f" matched, {counts['no_manual']} no manual data):")
    print(f"    ALL AGREE (green):         {counts['all_agree']}")
    print(f"    Rater_1 disagrees (yellow): {counts['r1_disagree']}")
    print(f"    Rater_2 disagrees (blue):   {counts['r2_disagree']}")
    print(f"    Manual disagrees (red):     {counts['manual_disagree']}")
    print(f"    All 3 disagree (gray):      {counts['all_disagree']}")


if __name__ == "__main__":
    main()
