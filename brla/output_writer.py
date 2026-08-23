"""Output writer — master.json + review.xlsx.

master.json: provenance-complete aggregation of all techs' final.json data,
with empty human_bin / human_notes / reviewed fields ready for fold-back.

review.xlsx: one row per tech-dimension, sorted for efficient human review
(needs_review desc, evidence_strength asc). Both raters' rationales shown
side-by-side with [N] / [Na] citation numbering. Per-source clickable columns.
"""
import re
import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .foldback import read_reviewed_xlsx
from .utils import read_json, tech_dir, write_json

FIXED_COLUMNS = [
    # Identity
    ("tech_id", 14),
    ("tech_name", 20),
    ("dimension", 10),
    # Assessment
    ("evidence_strength", 16),
    ("n_sources", 9),
    ("agreement", 10),
    ("evidence_gap", 12),
    ("juris_varies", 11),
    ("entity_conc", 18),
    ("single_rater", 12),
    ("needs_review", 12),
    # Human review
    ("human_bin", 10),
    ("human_notes", 30),
    ("reviewed", 9),
    # Rater 1
    ("rater_1_bin", 10),
    ("rater_1_level_est", 13),
    ("rater_1_rationale", 65),
    # Rater 2
    ("rater_2_bin", 10),
    ("rater_2_level_est", 13),
    ("rater_2_rationale", 55),
]

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4",
                          fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
REVIEW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC",
                          fill_type="solid")
LINK_FONT = Font(color="0563C1", underline="single", size=9)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
TOP_ALIGN = Alignment(vertical="top")



# ---------------------------------------------------------------------------
# Citation helpers
# ---------------------------------------------------------------------------

def _shorten_title(title: str, max_len: int = 40) -> str:
    if not title:
        return "source"
    title = re.sub(r"\s*\|.*", "", title)
    title = re.sub(r"^\[PDF\]\s*", "", title)
    if len(title) > max_len:
        title = title[:max_len - 1] + "…"
    return title


def _extract_eids_from_text(rationale: str, tech_id: str) -> list[str]:
    """Parse evidence_id references from rationale text, in first-seen order."""
    prefix = tech_id + "_"
    seen = set()
    result = []
    for m in re.finditer(
            r'(' + re.escape(tech_id) + r'_e\d+)|\b(e\d{3,})\b', rationale,
            re.IGNORECASE):
        if m.group(1):
            eid = m.group(1).lower()
        else:
            eid = prefix + m.group(2).lower()
        if eid not in seen:
            seen.add(eid)
            result.append(eid)
    return result


def _build_citation_map(p_rationale: str, s_rationale: str,
                        tech_id: str, evidence_by_id: dict):
    """Build unified citation numbering with sub-letters for same-URL evidence.

    Primary rater's references are numbered first, then second rater's new
    ones are appended.  Same URL = same [N]; multiple snippets from one URL
    get sub-letters [5a], [5b].

    Returns:
        eid_to_label: {evidence_id: "[5a]" or "[3]"}
        citation_list: [{num, url, title, snippets: [{letter, snippet}]}]
    """
    p_eids = _extract_eids_from_text(p_rationale, tech_id)
    s_eids = _extract_eids_from_text(s_rationale, tech_id)
    all_eids = list(dict.fromkeys(p_eids + s_eids))

    url_order: list[str] = []
    url_to_eids: dict[str, list[str]] = {}
    for eid in all_eids:
        rec = evidence_by_id.get(eid)
        if not rec:
            continue
        url = rec.get("url", "")
        if url not in url_to_eids:
            url_order.append(url)
            url_to_eids[url] = []
        url_to_eids[url].append(eid)

    eid_to_label: dict[str, str] = {}
    citation_list: list[dict] = []
    for num, url in enumerate(url_order, 1):
        eids = url_to_eids[url]
        rec0 = evidence_by_id[eids[0]]
        title = _shorten_title(rec0.get("title", ""))

        if len(eids) == 1:
            eid_to_label[eids[0]] = f"[{num}]"
            snippet = (evidence_by_id[eids[0]].get("snippet", "") or "")[:120]
            citation_list.append({
                "num": num, "url": url, "title": title,
                "snippets": [{"letter": None, "snippet": snippet}],
            })
        else:
            sub_snippets = []
            for i, eid in enumerate(eids):
                letter = chr(ord('a') + i)
                eid_to_label[eid] = f"[{num}{letter}]"
                snippet = (evidence_by_id[eid].get("snippet", "") or "")[:120]
                sub_snippets.append({"letter": letter, "snippet": snippet})
            citation_list.append({
                "num": num, "url": url, "title": title,
                "snippets": sub_snippets,
            })

    return eid_to_label, citation_list


def _replace_citations(rationale: str, tech_id: str,
                       eid_to_label: dict[str, str]) -> str:
    """Replace evidence_id tags in rationale text with [N] or [Na] labels."""
    def replace_full(m):
        return eid_to_label.get(m.group(0).lower(), m.group(0))

    def replace_short(m):
        full = tech_id + "_" + m.group(0).lower()
        return eid_to_label.get(full, m.group(0))

    result = re.sub(re.escape(tech_id) + r"_e\d+", replace_full, rationale,
                    flags=re.IGNORECASE)
    result = re.sub(r"\be\d{3,}\b", replace_short, result,
                    flags=re.IGNORECASE)
    return result


def _format_source_cell(citation: dict) -> str:
    lines = [f"[{citation['num']}] {citation['title']}"]
    for s in citation["snippets"]:
        if s["letter"]:
            lines.append(f"{s['letter']}) \"{s['snippet']}\"")
        else:
            lines.append(f"\"{s['snippet']}\"")
    return "\n".join(lines)


def _format_entity_conc(entry: dict) -> str:
    conc = entry.get("entity_concentration", 0.0)
    entity = entry.get("dominant_entity")
    if conc >= 0.05 and entity:
        return f"{entity} ({conc:.0%})"
    return ""


# ---------------------------------------------------------------------------
# Master JSON
# ---------------------------------------------------------------------------

def build_master(cfg: dict, techs: list[dict]) -> dict:
    """Aggregate all techs' final.json into a single master document."""
    entries = []
    for tech in techs:
        tdir = tech_dir(cfg, tech["tech_id"])
        final = read_json(tdir / "final.json")
        if not final:
            continue
        evidence = read_json(tdir / "evidence.json")
        all_records = evidence["records"] if evidence else []
        evidence_by_id = {r["evidence_id"]: r for r in all_records}

        for dim, d in final["dimensions"].items():
            dim_records = [r for r in all_records
                           if dim in r.get("dimensions", [])]

            p_rationale_raw = d.get("rationale", "")
            p_evidence_ids = d.get("evidence_ids", [])
            s_rationale_raw = d.get("second_rater", {}).get("rationale", "")
            s_evidence_ids = d.get("second_rater", {}).get("evidence_ids", [])

            entries.append({
                "tech_id": tech["tech_id"],
                "tech_name": tech["name"],
                "dimension": dim,
                "evidence_strength": d["evidence_strength"],
                "n_sources": d["n_evidence_records"],
                "rater_agreement": d["rater_agreement"],
                "evidence_gap": d["evidence_gap"],
                "jurisdiction_variation": d.get("jurisdiction_variation", False),
                "entity_concentration": d.get("entity_concentration", 0.0),
                "dominant_entity": d.get("dominant_entity"),
                "single_entity_risk": d.get("single_entity_risk", False),
                "needs_review": d["needs_review"],
                "degraded": d.get("degraded", False),
                "primary_rater": d["primary_rater"],
                "second_rater": d["second_rater"],
                "rater_1_rationale_raw": p_rationale_raw,
                "rater_1_evidence_ids": p_evidence_ids,
                "rater_2_rationale_raw": s_rationale_raw,
                "rater_2_evidence_ids": s_evidence_ids,
                "evidence_by_id": evidence_by_id,
                "human_bin": None,
                "human_notes": None,
                "reviewed": False,
            })

    entries.sort(key=lambda e: (-e["needs_review"], e["evidence_strength"]))
    return {"entries": entries, "n_techs": len(techs), "n_rows": len(entries)}


def write_master_json(cfg: dict, master: dict) -> Path:
    out_dir = cfg["paths"]["output_dir"]
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / "master.json"
    serializable = {
        "n_techs": master["n_techs"],
        "n_rows": master["n_rows"],
        "entries": [
            {k: v for k, v in e.items() if k != "evidence_by_id"}
            for e in master["entries"]
        ],
    }
    write_json(path, serializable)
    return path


def _carry_over_human_fields(path: Path) -> dict[tuple[str, str], dict]:
    """Read human review columns out of an existing review.xlsx."""
    if not path.exists():
        return {}
    try:
        return read_reviewed_xlsx(path)
    except Exception as e:  # noqa: BLE001
        print(f"  !! could not read existing {path.name} ({e}); "
              f"human review columns will be blank (see the .bak file)")
        return {}


# ---------------------------------------------------------------------------
# Review XLSX
# ---------------------------------------------------------------------------

def write_review_xlsx(cfg: dict, master: dict) -> Path:
    out_dir = cfg["paths"]["output_dir"]
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / "review.xlsx"

    carried = _carry_over_human_fields(path)
    if path.exists():
        shutil.copy2(path, path.with_suffix(".xlsx.bak"))

    # First pass: build citation data for every row, track max source columns
    row_data = []
    max_citations = 0
    for entry in master["entries"]:
        evidence_by_id = entry.get("evidence_by_id", {})
        tech_id = entry["tech_id"]

        eid_to_label, citation_list = _build_citation_map(
            entry.get("rater_1_rationale_raw", ""),
            entry.get("rater_2_rationale_raw", ""),
            tech_id, evidence_by_id)

        r1_rationale = _replace_citations(
            entry.get("rater_1_rationale_raw", ""), tech_id, eid_to_label)
        r2_rationale = _replace_citations(
            entry.get("rater_2_rationale_raw", ""), tech_id, eid_to_label)

        max_citations = max(max_citations, len(citation_list))

        human = carried.get((entry["tech_id"], entry["dimension"]), {})

        if entry["dimension"] == "RRL":
            juris = "Yes" if entry.get("jurisdiction_variation") else ""
        else:
            juris = ""

        row_data.append({
            "tech_id": entry["tech_id"],
            "tech_name": entry["tech_name"],
            "dimension": entry["dimension"],
            "evidence_strength": entry["evidence_strength"],
            "n_sources": entry["n_sources"],
            "agreement": "Yes" if entry["rater_agreement"] else "No",
            "evidence_gap": "Yes" if entry["evidence_gap"] else "",
            "juris_varies": juris,
            "entity_conc": _format_entity_conc(entry),
            "single_rater": "Yes" if entry.get("degraded") else "",
            "needs_review": "Yes" if entry["needs_review"] else "",
            "human_bin": human.get("human_bin") or "",
            "human_notes": human.get("human_notes") or "",
            "reviewed": "yes" if human.get("reviewed") else "",
            "rater_1_bin": entry["primary_rater"].get("bin") or "",
            "rater_1_level_est": entry["primary_rater"].get("level_estimate") or "",
            "rater_1_rationale": r1_rationale,
            "rater_2_bin": entry["second_rater"].get("bin") or "",
            "rater_2_level_est": entry["second_rater"].get("level_estimate") or "",
            "rater_2_rationale": r2_rationale,
            "citations": citation_list,
        })

    # Build column list: fixed + dynamic source columns
    source_cols = [(f"source_{i}", 35) for i in range(1, max_citations + 1)]
    all_cols = list(FIXED_COLUMNS) + source_cols

    rationale_col_indices = set()
    for i, (name, _) in enumerate(FIXED_COLUMNS, 1):
        if "rationale" in name:
            rationale_col_indices.add(i)

    wb = Workbook()
    ws = wb.active
    ws.title = "BRLa Review"
    ws.freeze_panes = "A2"

    # Headers
    for col_idx, (name, width) in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = TOP_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    fixed_col_names = [name for name, _ in FIXED_COLUMNS]

    # Data rows
    for row_idx, row in enumerate(row_data, 2):
        # Fixed columns
        for col_idx, col_name in enumerate(fixed_col_names, 1):
            val = row[col_name]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx in rationale_col_indices:
                cell.alignment = WRAP_ALIGN
            else:
                cell.alignment = TOP_ALIGN

        # Highlight needs_review rows
        if row["needs_review"] == "Yes":
            for col_idx in range(1, len(FIXED_COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = REVIEW_FILL

        # Source columns
        source_start = len(FIXED_COLUMNS) + 1
        for cit in row["citations"]:
            col_idx = source_start + cit["num"] - 1
            label = _format_source_cell(cit)
            cell = ws.cell(row=row_idx, column=col_idx, value=label)
            cell.font = LINK_FONT
            cell.alignment = WRAP_ALIGN
            if cit["url"].startswith("http"):
                cell.hyperlink = cit["url"]

    # Group source columns (collapsible in Excel)
    if max_citations > 0:
        source_start_letter = get_column_letter(len(FIXED_COLUMNS) + 1)
        source_end_letter = get_column_letter(len(FIXED_COLUMNS) + max_citations)
        ws.column_dimensions.group(source_start_letter, source_end_letter,
                                   hidden=False)

    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(all_cols))}{len(row_data) + 1}")
    wb.save(path)
    return path


def run(cfg: dict, techs: list[dict]) -> dict:
    master = build_master(cfg, techs)
    json_path = write_master_json(cfg, master)
    xlsx_path = write_review_xlsx(cfg, master)
    return {
        "n_rows": master["n_rows"],
        "n_techs": master["n_techs"],
        "master_json": str(json_path),
        "review_xlsx": str(xlsx_path),
    }
